import openpyxl

categories = ((0,""),(10,"폴로 셔츠"), (11,"드레스 셔츠"), (12,"캐주얼 셔츠"), \
            (13,"티셔츠"), (15,"가디건"),(16,"스웨터"),(17,"땀받이 셔츠"), \
            (18,"파카"))
sizes = ("코드","분류명","S","M","L","LL","XL")
#2차원 리스트 작성
# 요소 리스트가 모두 같은 객체가 되므로 주석 처리된 코드는 사용 불가
#order_amount= [[0]*len(sizes)] * len(categories)
order_amount= [[0]*len(sizes) for i in range(len(categories))]

for j in range(len(sizes)):
    order_amount[0][j] = sizes[j]

for i in range(1,len(categories)):
    order_amount[i][0] = categories[i][0]
    order_amount[i][1] = categories[i][1]
#print(order_amount)   

wb = openpyxl.load_workbook("..\data\ordersList.xlsx")
sh = wb.active
for row in range(2, sh.max_row + 1):
    category = sh["I" + str(row)].value
    size = sh["L" + str(row)].value
    amount = sh["M" + str(row)].value
    for i in range(1,len(categories)):
        if category == order_amount[i][0]:
            for j in range(2,len(sizes)):
                if size == order_amount[0][j]:
                    order_amount[i][j] += amount


owb = openpyxl.Workbook()
osh = owb.active
row = 1
for order_row in order_amount:
    col = 1
    size_sum = 0 
    for order_col in order_row:
        osh.cell(row, col).value = order_col
        if  row > 1 and col > 2: 
            #print(order_col)
            size_sum += order_col
        col += 1
    if row == 1:
        osh.cell(row, col).value = "합계"
    else:
        osh.cell(row, col).value = size_sum
    row += 1

owb.save("..\data\orders_aggregate.xlsx")