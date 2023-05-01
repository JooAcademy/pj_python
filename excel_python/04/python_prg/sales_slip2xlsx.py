import pathlib
import openpyxl
import csv


lwb = openpyxl.Workbook()
lsh = lwb.active
list_row = 1
path = pathlib.Path("..\data\sales")
for pass_obj in sorted(path.iterdir()):
    if pass_obj.match("*.xlsx"):
        wb = openpyxl.load_workbook(pass_obj)
        for sh in wb:
            for dt_row in range(9,19):
                if sh.cell(dt_row, 2).value != None:
                    # 보다 구체적인 코드
                    #lsh.cell(row=list_row, column=1).value = \
                    #    sh.cell(row=2, column=7).value   #전표NO
                    lsh.cell(list_row, 1).value = sh.cell(2, 7).value   #전표NO
                    lsh.cell(list_row, 2).value = sh.cell(3, 7).value   #일시
                    lsh.cell(list_row, 3).value = sh.cell(4, 3).value   #거래처 코드
                    lsh.cell(list_row, 4).value = sh.cell(3, 2).value.strip(" 귀하")   #거래처명
                    lsh.cell(list_row, 5).value = sh.cell(7, 8).value   #담당자 코드
                    lsh.cell(list_row, 6).value = sh.cell(7, 7).value   #담당자명
                    lsh.cell(list_row, 7).value = sh.cell(dt_row, 1).value #No
                    lsh.cell(list_row, 8).value = sh.cell(dt_row, 2).value #상품 코드
                    lsh.cell(list_row, 9).value = sh.cell(dt_row, 3).value #상품명
                    lsh.cell(list_row, 10).value = sh.cell(dt_row, 4).value #수량
                    lsh.cell(list_row, 11).value = sh.cell(dt_row, 5).value #단가
                    lsh.cell(list_row, 12).value = sh.cell(dt_row, 4).value * \
                                                sh.cell(dt_row, 5).value #금액
                    lsh.cell(list_row, 13).value = sh.cell(dt_row, 7).value #비고
                    list_row += 1

lwb.save("..\data\salesList.xlsx")
