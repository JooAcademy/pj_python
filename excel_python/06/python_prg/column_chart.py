import openpyxl
from openpyxl.chart import BarChart, Reference

wb = openpyxl.load_workbook("..\data\column_chart.xlsx")
sh = wb.active
#print(sh.max_row)
data = Reference(sh, min_col=3, max_col=3, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=2, max_col=2, min_row=2, max_row=sh.max_row)
chart = BarChart()
chart.type = "col"
#chart.type = "bar" #가로 막대형
# 가로 막대형이 되면 거래처 일부가 사라지므로 전체 높이를 지정
#chart.height = 10

chart.style = 28    #1 회색, 11 파란색, 28 오렌지색, 30 노란색, 37 연한 회색 배경, 45 검은 배경
chart.title = "거래처별 매출"
chart.y_axis.title = "매출액"
chart.x_axis.title = "거래처명"

chart.add_data(data,titles_from_data=True)  #월매출이 범례가 됨
chart.set_categories(labels)
sh.add_chart(chart, "E3")

wb.save("..\data\column_chart.xlsx")