import openpyxl
from openpyxl.chart import AreaChart, Reference

wb = openpyxl.load_workbook(r"..\data\area_chart.xlsx")
sh = wb.active

data = Reference(sh, min_col=3, max_col=7, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=2, max_col=2, min_row=2, max_row=sh.max_row)
chart = AreaChart()
chart.grouping = "stacked"
#chart.grouping = "percentStacked"
chart.title = "상품 분류별 판매량(사이즈 누적)"
chart.x_axis.title = "분류"
chart.y_axis.title = "사이즈"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sh.add_chart(chart, "I2")
wb.save(r"..\data\area_chart.xlsx")