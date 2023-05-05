import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

#상수
TITLE_CELL_COLOR = "AA8866"

wb = openpyxl.load_workbook("..\data\orders_aggregate.xlsx")
sh = wb.active

#1행, 2열 고정
sh.freeze_panes = "C2"
#열 크기 지정
col_widths = {"A":8, "B":15, "C":10, "D":10, "E":10, \
    "F":10, "G":10, "H":10}
for col_name in col_widths:
    sh.column_dimensions[col_name].width = col_widths[col_name]

#A열 숨기기
#sh.column_dimensions["A"].hidden=False

for i in range(2, sh.max_row+1):
    sh.row_dimensions[i].height = 18
    for j in range(3, sh.max_column+1):
        #자릿수 표시
        sh.cell(row=i,column=j).number_format = "#,##0" 
        if j == 8:
            sh.cell(row=i,column=j).font = Font(bold=True)

#폰트 지정
font_header = Font(name="맑은 고딕",size=12,bold=True,color="FFFFFF")

for rows in sh["A1":"H1"]:
    for cell in rows:
        #셀 배경색
        cell.fill = PatternFill(patternType="solid", fgColor=TITLE_CELL_COLOR)
        #가로 위치 균등 분할
        cell.alignment = Alignment(horizontal="distributed")
        #폰트 지정
        cell.font = font_header

side = Side(style="thin", color="000000")
border = Border(left=side, right=side, top=side, bottom=side)
for row in sh:
    for cell in row:
        cell.border = border
        #sh[cell.coordinate].border = border

wb.save("..\data\orders_aggregate_ed.xlsx")
