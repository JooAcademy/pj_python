import pathlib
import openpyxl
import csv

lwb = openpyxl.Workbook()
lsh = lwb.active
list_row = 1
path = pathlib.Path("..\data\sales")
for pass_obj in path.iterdir(): # 현재 폴더와 하위 폴더의 파일까지 읽어서 반환
    if pass_obj.match("*.xlsx"):
        wb = openpyxl.load_workbook(pass_obj)
        for sh in wb:
            for dt_row in range(9,19):
                if sh.cell(dt_row, 2).value != None:
                    lsh.cell(list_row, 1).value = sh.cell(2, 7).value # 2행 7열 전표No
                    lsh.cell(list_row, 2).value = sh.cell(3, 7).value # 3행 7열 일시
                    lsh.cell(list_row, 3).value = sh.cell(4, 3).value # 4행 3열 거래처코드
                    lsh.cell(list_row, 4).value = sh.cell(7, 8).value # 7행 8열 담당자코드
                    
                    lsh.cell(list_row, 5).value = sh.cell(dt_row, 1).value # 9행 1열 순번                    
                    lsh.cell(list_row, 6).value = sh.cell(dt_row, 2).value # 9행 2열 상품코드
                    lsh.cell(list_row, 7).value = sh.cell(dt_row, 3).value # 9행 3열 품명
                    lsh.cell(list_row, 8).value = sh.cell(dt_row, 4).value # 9행 4열 수량
                    lsh.cell(list_row, 9).value = sh.cell(dt_row, 5).value # 9행 5열 단가
                    lsh.cell(list_row, 10).value = sh.cell(dt_row, 4).value * sh.cell(dt_row, 5).value # 9행 6열 금액
                    lsh.cell(list_row, 11).value = sh.cell(dt_row, 7).value # 9행 7열 비고
                    list_row += 1
#lwb.save("..\data\sales\salesList.xlsx") # 엑셀로 저장 모드
with open("..\data\sales\salesList.csv","w",encoding="utf_8_sig") as fp: # 중간 "w"가 cvs 쓰기 모드
    writer = csv.writer(fp, lineterminator="\n")
    for row in lsh.rows:
        writer.writerow([col.value for col in row])
        print([col.value for col in row])
