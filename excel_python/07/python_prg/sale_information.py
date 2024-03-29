from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, portrait
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.lib.units import cm
import openpyxl 
import pathlib  
import datetime
from PIL import Image

def load_information():
    wb = openpyxl.load_workbook("..\data\세일안내.xlsx")
    sh = wb.active
    sale_dict = {}
    for row in range(1, sh.max_row + 1):
        if sh.cell(row,1).value == "안내문":
            info_list = [sh.cell(row,2).value]
            for info_row in range(row + 1 , sh.max_row + 1):
                info_list.append(sh.cell(info_row,2).value)
            sale_dict.setdefault("안내문", info_list)
        elif sh.cell(row,1).value is not None:     
            sale_dict.setdefault(sh.cell(row,1).value, sh.cell(row,2).value)
    return sale_dict


sale_dict = load_information()
path = pathlib.Path("..\data\sales\pdf")
wb = openpyxl.load_workbook("..\data\거래처장부.xlsx")
sh = wb["주소록"]
for row in range(1, sh.max_row + 1):
    file_name = (sh.cell(row,2).value) + "님 안내문.pdf"
    out_path =  path / file_name
    cv = canvas.Canvas(str(out_path), pagesize=portrait(A4))
    cv.setTitle("세일 안내")
    pdfmetrics.registerFont(UnicodeCIDFont("HYGothic-Medium"))
    cv.setFont("HYGothic-Medium", 12)
    cv.drawCentredString(6*cm, 27*cm, sh.cell(row,2).value + " " \
        + sh.cell(row,3).value + " 귀하")
    cv.line(1.8*cm, 26.8*cm,10.8*cm,26.8*cm)
    cv.setFont("HYGothic-Medium", 14)
    cv.drawCentredString(10*cm, 24*cm, sale_dict["제목"])
    cv.setFont("HYGothic-Medium", 12)
    cv.drawString(2*cm, 22*cm, "개최일시: " + sale_dict["개최일시"])
    cv.drawString(2*cm, 21*cm, "개최장소: " + sale_dict["개최장소"])

    textobject = cv.beginText()
    textobject.setTextOrigin(2*cm, 19*cm,)
    textobject.setFont("HYGothic-Medium", 12)
    for line in sale_dict["안내문"]:
        textobject.textOut(line)
        textobject.moveCursor(0,14)
    
    cv.drawText(textobject)
    now = datetime.datetime.now()
    cv.drawString(14.4*cm, 14.8*cm, now.strftime("%Y/%m/%d"))
    image =Image.open("..\data\logo.png")
    cv.drawInlineImage(image,12*cm,11*cm)
    cv.showPage()
    cv.save()
